"""
Script para generar labels.csv desde la base de datos.
Este archivo contiene datos históricos con etiquetas "ground truth".
INCLUYE: Exportación a Excel con formato condicional
"""
import pandas as pd
import psycopg2
from datetime import datetime
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

# Configuración de base de datos
DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'greendelivery',
    'user': 'postgres',
    'password': '1234'
}

# Umbrales (iguales al detector)
UMBRAL_TEMPERATURA = 8.0
UMBRAL_FUERZA_G = 2.5
UMBRAL_INCLINACION = 30.0
UMBRAL_VIBRACION = 4.0  # Nuevo umbral para vibración

def conectar_bd():
    """Conectar a PostgreSQL"""
    return psycopg2.connect(**DB_CONFIG)

def generar_labels():
    """
    Genera el archivo labels.csv con datos etiquetados.
    
    ESTRATEGIA MEJORADA:
    1. Leer telemetría y alertas
    2. Identificar TODOS los eventos anómalos (no solo los de alertas)
    3. Etiquetar según:
       - Incidente confirmado (parte de una alerta)
       - Evento anómalo aislado (supera umbral pero no generó alerta)
       - Normal
    """
    print("🔄 Conectando a la base de datos...")
    conn = conectar_bd()
    
    # Leer telemetría
    print("📊 Leyendo datos de telemetría...")
    df_telemetry = pd.read_sql("""
        SELECT 
            id,
            id_paquete,
            timestamp,
            temperatura,
            fuerza_g,
            inclinacion,
            humedad,
            oxigeno,
            vapores,
            iluminacion,
            vibracion
        FROM telemetry
        ORDER BY id
    """, conn)
    
    print(f"   └─ {len(df_telemetry)} eventos de telemetría cargados")
    
    # Leer alertas
    print("🚨 Leyendo alertas...")
    df_alerts = pd.read_sql("""
        SELECT 
            id,
            id_paquete,
            tipo_incidente,
            timestamp_inicio,
            timestamp_fin,
            num_eventos
        FROM alerts
        ORDER BY id
    """, conn)
    
    print(f"   └─ {len(df_alerts)} alertas cargadas")
    
    conn.close()
    
    # Crear columnas de etiqueta
    df_telemetry['incidente'] = 0
    df_telemetry['tipo_incidente'] = 'normal'
    df_telemetry['supera_umbral_temp'] = df_telemetry['temperatura'] > UMBRAL_TEMPERATURA
    df_telemetry['supera_umbral_choque'] = (
        (df_telemetry['fuerza_g'] > UMBRAL_FUERZA_G) & 
        (df_telemetry['inclinacion'] > UMBRAL_INCLINACION)
    )
    df_telemetry['supera_umbral_vibracion'] = df_telemetry['vibracion'] > UMBRAL_VIBRACION
    df_telemetry['pico_temperatura_extremo'] = df_telemetry['temperatura'] > 20.0  # > 20°C es extremo
    
    # Convertir timestamps
    df_telemetry['timestamp_dt'] = pd.to_datetime(df_telemetry['timestamp'])
    
    # ==========================================
    # ETIQUETAR EVENTOS DENTRO DE ALERTAS
    # ==========================================
    for idx, alert in df_alerts.iterrows():
        print(f"\n🔍 Procesando alerta {alert['id']}: {alert['tipo_incidente']}")
        
        ts_inicio = pd.to_datetime(alert['timestamp_inicio'])
        ts_fin = pd.to_datetime(alert['timestamp_fin']) if pd.notna(alert['timestamp_fin']) else None
        
        # Filtrar eventos de este paquete
        mask_paquete = df_telemetry['id_paquete'] == alert['id_paquete']
        mask_tiempo = df_telemetry['timestamp_dt'] >= ts_inicio
        
        if ts_fin:
            mask_tiempo = mask_tiempo & (df_telemetry['timestamp_dt'] <= ts_fin)
        else:
            # Si no hay fin, tomar los N eventos desde el inicio
            eventos_paquete = df_telemetry[mask_paquete & mask_tiempo].head(alert['num_eventos'])
            mask_tiempo = df_telemetry['id'].isin(eventos_paquete['id'])
        
        mask_final = mask_paquete & mask_tiempo
        num_etiquetados = mask_final.sum()
        
        # Aplicar etiquetas
        df_telemetry.loc[mask_final, 'incidente'] = 1
        df_telemetry.loc[mask_final, 'tipo_incidente'] = alert['tipo_incidente']
        
        print(f"   └─ {num_etiquetados} eventos etiquetados como '{alert['tipo_incidente']}'")
    
    # ==========================================
    # IDENTIFICAR EVENTOS ANÓMALOS NO ETIQUETADOS
    # ==========================================
    # Picos de temperatura extremos (FALSO POSITIVO - azul)
    mask_pico_extremo = (
        (df_telemetry['pico_temperatura_extremo']) & 
        (df_telemetry['incidente'] == 0)
    )
    df_telemetry.loc[mask_pico_extremo, 'tipo_incidente'] = 'pico_temperatura_extremo'
    
    # Picos de temperatura normales (amarillo)
    mask_pico_temp = (
        (df_telemetry['supera_umbral_temp']) & 
        (df_telemetry['incidente'] == 0) &
        (~df_telemetry['pico_temperatura_extremo'])
    )
    df_telemetry.loc[mask_pico_temp, 'tipo_incidente'] = 'pico_temperatura'
    
    # Picos de fuerza_g (amarillo)
    mask_pico_choque = (
        (df_telemetry['supera_umbral_choque']) & 
        (df_telemetry['incidente'] == 0)
    )
    df_telemetry.loc[mask_pico_choque, 'tipo_incidente'] = 'pico_fuerza_g'
    
    # Vibraciones altas (amarillo)
    mask_vibracion_alta = (
        (df_telemetry['supera_umbral_vibracion']) & 
        (df_telemetry['incidente'] == 0)
    )
    df_telemetry.loc[mask_vibracion_alta, 'tipo_incidente'] = 'vibracion_alta'
    
    # ==========================================
    # ESTADÍSTICAS
    # ==========================================
    print("\n" + "="*60)
    print("📈 ESTADÍSTICAS DEL DATASET:")
    print("="*60)
    print(f"Total de eventos: {len(df_telemetry)}")
    print(f"\nPor categoría:")
    
    eventos_normales = (df_telemetry['tipo_incidente'] == 'normal').sum()
    eventos_alertas = (df_telemetry['incidente'] == 1).sum()
    eventos_pico_temp = (df_telemetry['tipo_incidente'] == 'pico_temperatura').sum()
    eventos_pico_choque = (df_telemetry['tipo_incidente'] == 'pico_fuerza_g').sum()
    eventos_pico_extremo = (df_telemetry['tipo_incidente'] == 'pico_temperatura_extremo').sum()
    eventos_vibracion = (df_telemetry['tipo_incidente'] == 'vibracion_alta').sum()
    
    print(f"  • Normales: {eventos_normales} ({eventos_normales / len(df_telemetry) * 100:.1f}%)")
    print(f"  • Incidentes (alertas): {eventos_alertas} ({eventos_alertas / len(df_telemetry) * 100:.1f}%)")
    print(f"  • Picos temperatura: {eventos_pico_temp}")
    print(f"  • Picos fuerza_g: {eventos_pico_choque}")
    print(f"  • Picos temperatura extremos: {eventos_pico_extremo} 🔵")
    print(f"  • Vibraciones altas: {eventos_vibracion} 🟡")
    
    print("\n📊 Distribución por tipo de incidente:")
    print(df_telemetry['tipo_incidente'].value_counts())
    print("="*60)
    
    # ==========================================
    # GUARDAR CSV
    # ==========================================
    output_file = 'analytics/labels.csv'
    os.makedirs('analytics', exist_ok=True)
    
    # Eliminar columnas auxiliares antes de guardar
    df_export = df_telemetry.drop([
        'timestamp_dt', 'supera_umbral_temp', 'supera_umbral_choque',
        'supera_umbral_vibracion', 'pico_temperatura_extremo'
    ], axis=1)
    df_export.to_csv(output_file, index=False)
    
    print(f"\n✅ CSV guardado: {output_file}")
    
    # ==========================================
    # GENERAR EXCEL CON FORMATO CONDICIONAL
    # ==========================================
    print("\n📊 Generando Excel con formato condicional...")
    
    excel_file = 'analytics/labels_formatted.xlsx'
    
    # Crear Excel
    df_export.to_excel(excel_file, index=False, sheet_name='Telemetry Data')
    
    # Abrir el archivo para aplicar formato
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Colores VIVOS (mucho más visibles)
    color_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    color_amarillo = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")  # Amarillo fuerte
    color_rojo = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo intenso
    color_azul = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")  # Azul turquesa llamativo
    
    font_bold = Font(bold=True)
    
    # Aplicar formato a las filas (empezamos en 2 porque 1 es el header)
    for row_idx, row in enumerate(df_export.itertuples(), start=2):
        tipo = row.tipo_incidente
        incidente = row.incidente
        
        # Determinar color
        if incidente == 1:
            # Parte de una alerta confirmada → ROJO
            fill = color_rojo
            bold = True
        elif tipo == 'pico_temperatura_extremo':
            # Pico extremo (falso positivo) → AZUL
            fill = color_azul
            bold = False
        elif tipo in ['pico_temperatura', 'pico_fuerza_g', 'vibracion_alta']:
            # Picos aislados o vibraciones → AMARILLO
            fill = color_amarillo
            bold = False
        else:
            # Normal → BLANCO
            fill = color_normal
            bold = False
        
        # Aplicar formato a toda la fila
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            if bold:
                cell.font = font_bold
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila (headers)
    ws.freeze_panes = 'A2'
    
    # Guardar
    wb.save(excel_file)
    print(f"✅ Excel con formato guardado: {excel_file}")
    print("\n📌 LEYENDA DE COLORES:")
    print("  🔴 ROJO: Eventos que forman parte de una ALERTA confirmada (3+ eventos consecutivos)")
    print("  🔵 AZUL: Pico de temperatura EXTREMO (>20°C) - Falso positivo")
    print("  🟡 AMARILLO: Picos aislados o vibraciones altas - Eventos a monitorizar")
    print("  ⚪ BLANCO: Eventos normales")
    
    return df_export


if __name__ == "__main__":
    print("="*60)
    print("🏷️  GENERADOR DE LABELS.CSV")
    print("="*60)
    df = generar_labels()
    print("\n✅ ¡Proceso completado!")
    print("\nArchivos generados:")
    print("  • analytics/labels.csv (para scripts Python)")
    print("  • analytics/labels_formatted.xlsx (para revisar en Excel)")